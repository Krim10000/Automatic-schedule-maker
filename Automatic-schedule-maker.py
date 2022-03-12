
import numpy as np

"""
Profesores =["Dulce", "Linda"]

print (Profesores)


1.-Bloque horario :)

2.-Requerimiento de curso

3.-Profesor



Bloque horario:

Día
N° de bloque
curso
"""

#Crea los Bloques de tiempo a ser asignados
######################################################################
Dias = ["Lunes", "Martes", "Miercoles", "Jueves", "Viernes"] # lista
Bloques= 8 # numero
#Cursos = ["I°A","II°A","III°A","IV°A"] #lista

Cursos = ["7°A","7°B","8°A","8°B"]

#para 7 y 8

Lista_BLOQUES=[]
contador_bloques=1

       
for D in Dias:
        for C in Cursos:
                if D== "Lunes":
                        for B in range(1,8+1):
                                BLOQUE= (D,C,B)
                                Lista_BLOQUES.append(BLOQUE)
                elif D== "Viernes":
                
                        for B in range(1,6+1):
                                BLOQUE= (D,C,B)
                                Lista_BLOQUES.append(BLOQUE)
                else:
                        for B in range(1,Bloques+1):
                                BLOQUE= (D,C,B)
                                Lista_BLOQUES.append(BLOQUE)

#print(len(Lista_BLOQUES))

print("Total de bloques: "+str(len(Lista_BLOQUES)))
print("Bloques por curso: "+str(len(Lista_BLOQUES)/len(Cursos)))

#############################################################################

# 37 bloques por semana media
# 38 bloque para 7 y 8

"""
para 7 y 8

musica o arte	3
edu fis	2
ciencia	4
matematica	8
len 	8
ingles	3
Historia	4
tecnologia 	1
orientacion 	1
F valorica	2
F ciudaddana	2

"""
#Requerimientos de septimo y octavo
#########################

Ramos = [("Musica o Artes",3),("Educación Fisica",2),("Ciencias",4),("Matematica",8),("Lenguaje",8), ("Ingles",3),
                  ("Historia",4), ("Tecnología",1), ("Orientación",1),("Formación valorica",2),("Formación ciudadana",2) ]



Total_Req = 0
for k in range(len(Ramos)):
        Total_Req = Total_Req + int(Ramos[k][1])

print("Total de requerimientos: "+str(Total_Req))




#######################

#Profesores

#profesor, ramo, curso

Profesores = [

("Mozart","Musica o Artes","7°A"),("Rambo","Educación Fisica","7°A"),("Einsteins","Ciencias","7°A"),
("Dirac","Matematica","7°A",),("Cervantes","Lenguaje","7°A"), ("Shakespeare","Ingles","7°A"),
("Origenes","Historia","7°A"),("Newton","Tecnología","7°A"),("Guru","Orientación","7°A"),
("Aquino","Formación valorica","7°A"),("Portales","Formación ciudadana","7°A"),


("Mozart","Musica o Artes","7°B"),("Rambo","Educación Fisica","7°B"),("Einsteins","Ciencias","7°B"),
("Dirac","Matematica","7°B",),("Cervantes","Lenguaje","7°B"), ("Shakespeare","Ingles","7°B"),
("Origenes","Historia","7°B"),("Newton","Tecnología","7°B"),("Guru","Orientación","7°B"),
("Aquino","Formación valorica","7°B"),("Portales","Formación ciudadana","7°B"),


("Mozart","Musica o Artes","8°A"),("Rambo","Educación Fisica","8°A"),("Einsteins","Ciencias","8°A"),
("Dirac","Matematica","8°A",),("Cervantes","Lenguaje","8°A"), ("Shakespeare","Ingles","8°A"),
("Origenes","Historia","8°A"),("Newton","Tecnología","8°A"),("Guru","Orientación","8°A"),
("Aquino","Formación valorica","8°A"),("Portales","Formación ciudadana","8°A"),


("Mozart","Musica o Artes","8°B"),("Rambo","Educación Fisica","8°B"),("Einsteins","Ciencias","8°B"),
("Dirac","Matematica","8°B",),("Cervantes","Lenguaje","8°B"), ("Shakespeare","Ingles","8°B"),
("Origenes","Historia","8°B"),("Newton","Tecnología","8°B"),("Guru","Orientación","8°B"),
("Aquino","Formación valorica","8°B"),("Portales","Formación ciudadana","8°B")

]
                            
              
                        
Lista_profesores=[]                
Total_prof = 0
for l in range(len(Profesores)):

        profe=Profesores[l][0]
        Lista_profesores.append(profe)
        

#print("Total de profesores "+str(Total_prof)) 

print ("Número de profesores: "+ str(len(set(Lista_profesores))))

        
########################
                
 #Lista_BLOQUES --> Ramos --> Profesores




#####################


#Determinación
#BLOQUE, Ramo, profesor.



#########################################################

#Ramos por curso


Ramos_curso=[]
#Ramos
#Cursos

for c in range(len(Cursos)):
        for r in range(len(Ramos)):
                for x in range(int(Ramos[r][1])):
               
                        mau=Ramos[r][0],Cursos[c],"Contador "+ str(x+1)
                        Ramos_curso.append(mau)
#########################################################

import random

def mezclar_lista(lista_original):
    # Crear una copia, ya que no deberíamos modificar la original
    # https://parzibyte.me/blog/2020/05/31/python-clonar-lista-eliminar-referencia/
    lista = lista_original[:]
    # Ciclo for desde 0 hasta la longitud de la lista -1
    longitud_lista = len(lista)
    for i in range(longitud_lista):
        # Obtener un índice aleatorio
        # https://parzibyte.me/blog/2019/04/04/generar-numero-aleatorio-rango-python/
        indice_aleatorio = random.randint(0, longitud_lista - 1)
        # Intercambiar
        temporal = lista[i]
        lista[i] = lista[indice_aleatorio]
        lista[indice_aleatorio] = temporal
    # Regresarla
    return lista
                
R_Lista_BLOQUES =[]

#########################################################
def ordenar():
        
        lista_temp=[]
        #Ramos_curso= np.array(Ramos_curso)
        #Ramos_curso esta mal configurado !!!!
        global Def
        Def=[]
        
        global R_Lista_BLOQUES
        R_Lista_BLOQUES =[]
        R_Ramos_curso =[]

        Restriccion =[]

        count=0


        global Lista_BLOQUES

        Lista_BLOQUES=mezclar_lista(Lista_BLOQUES) # orden aleatorio


        while len(Lista_BLOQUES)!=len(R_Lista_BLOQUES):
                if count == 5:
                        print("solo se asignaron " + str(len(R_Lista_BLOQUES))+ "/"+str(len(Lista_BLOQUES)) +" bloques"+ " luego de " + str(count) + " intentos")
                        ordenar()
                        break
                        
                count=count+1
                for x in range(len(Lista_BLOQUES)):

                        for y in range(len(Ramos_curso)):
                                if Lista_BLOQUES[x] in R_Lista_BLOQUES or Ramos_curso[y] in R_Ramos_curso: #Funciona :)

                                        continue
                            
                                for p in range(len(Profesores)):
                                        
                                        if Ramos_curso[y][0] ==Profesores[p][1] and Ramos_curso[y][1]==Profesores[p][2] :
                                                #calza un profesor

                                                
                                                AUX2 =str(Lista_BLOQUES[x][0])+str(Lista_BLOQUES[x][2])+str(Profesores[p][0])
                                                                                                            

                                                if AUX2 in Restriccion :
                                                                                                            
                                                       continue
                                                
                                                                                                            
                                                AUX = Lista_BLOQUES[x],Ramos_curso[y],Profesores[p]
                                                Def.append(AUX)
                                                R_Lista_BLOQUES.append(Lista_BLOQUES[x])
                                                R_Ramos_curso.append(Ramos_curso[y])
                                                Restriccion.append(AUX2)
                                                
        print(len(Def))                                                                                     

                #Ramos_curso[y][0] ramo
                #Ramos_curso[y][1] curso
                #Profesores[][1] ramo
                #Profesores[][2] curso
                #Lista_BLOQUES[x][0] dia
                #Lista_BLOQUES[x][2] hora
                                 
if len(Lista_BLOQUES)!=len(R_Lista_BLOQUES):
        ordenar()
print(len(Def))                 
########################################################################################################################
                
Def2=[] #(('Martes', '8°B', 2), 'Musica o Artes', 'Mozart')


for d in range(len(Def)):
        AUX3= Def[d][0],Def[d][1][0],Def[d][2][0]
        Def2.append(AUX3)



#########################################################################################################################

#Exportar horario a Excel

import re
import pandas as pd
import openpyxl
from openpyxl import Workbook
import datetime 

start_time = datetime.datetime.now()

#filename of the excel file
name ='Horario.xlsx'


wb = Workbook()
ws =  wb.active

#Hoja del archivo

for c in range(len(Cursos)):
        aux = "ws"+str(c)
        aux= wb.create_sheet(Cursos[c])

        
        #Def2=[] (('Martes', '8°B', 2), 'Musica o Artes', 'Mozart')

        for d in range(len(Def2)):
                if Cursos[c] != Def2[d][0][1]: #cursos
                        continue

                for b in range((Bloques)):
                        
                        for dd in range(len(Dias)):

                                aux.cell(row=1, column= dd+2).value=Dias[dd] #dias
                                aux.cell(row=b+2, column=1).value=b+1 #bloque
                                if Dias[dd] == Def2[d][0][0] and Def2[d][0][2]== b+1:
                                        aux.cell(row=b+2, column=dd+2).value= Def2[d][1]+" "+Def2[d][2] #ramo-profesor
                                        

########################################################################################################################                        
                
                        
                        


        
        


wb.save(filename = name)
print("Listo")




        
                                      
  
